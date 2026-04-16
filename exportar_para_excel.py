#!/usr/bin/env python3
"""
exportar_para_excel.py
══════════════════════════════════════════════════════════════════════════════
Extrai dados de navios de diesel dos portos brasileiros monitorados e
gera um relatório em Excel (.xlsx) com formatação profissional.

Como usar:
    python exportar_para_excel.py

Dependências (já listadas em requirements.txt):
    pip install requests beautifulsoup4 pandas openpyxl lxml selenium

O arquivo será salvo na mesma pasta do script com o nome:
    Navios_Diesel_AAAAMMDD_HHMM.xlsx
══════════════════════════════════════════════════════════════════════════════
"""

import sys
import traceback
from datetime import datetime, timezone, timedelta
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Importar funções de coleta ─────────────────────────────────────────────
try:
    from navios_esperados import (
        buscar_santos_esperados,
        buscar_santos_atracados,
        buscar_itaqui,
        buscar_paranagua,
        buscar_sao_sebastiao,
        buscar_suape,
        consolidar,
    )
except ImportError:
    print("╔══════════════════════════════════════════════════════════════╗")
    print("║  ERRO: navios_esperados.py não encontrado na pasta atual.    ║")
    print("║  Coloque este script na mesma pasta que navios_esperados.py  ║")
    print("╚══════════════════════════════════════════════════════════════╝")
    sys.exit(1)


# ── Paleta de cores (mesma do dashboard) ───────────────────────────────────
C_ORANGE   = "FF5000"
C_NAVY     = "000512"
C_WHITE    = "FFFFFF"
C_LIGHT_OG = "FFF3ED"
C_ALT_ROW  = "F7F7F7"
C_BORDER   = "CCCCCC"
C_SUCCESS  = "D4EDDA"   # verde claro  → Atracado
C_WARN     = "FFF3CD"   # amarelo       → Esperado
C_INFO     = "D1ECF1"   # azul claro   → Ao Largo / Fundeado
C_MUTED    = "E2E3E5"   # cinza         → Despachado
C_ERROR    = "F8D7DA"   # vermelho      → ERRO_COLETA / falha
C_OK_DARK  = "155724"   # texto verde escuro
C_ERR_DARK = "721C24"   # texto vermelho escuro

STATUS_FILL = {
    "Atracado":   C_SUCCESS,
    "Fundeado":   C_INFO,
    "Ao Largo":   C_INFO,
    "Esperado":   C_WARN,
    "Programado": C_WARN,
    "Despachado": C_MUTED,
    "ERRO_COLETA": C_ERROR,
}

# ── Helpers de estilo ──────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def _font(size=10, bold=False, color=C_NAVY, italic=False) -> Font:
    return Font(name="Arial", size=size, bold=bold, color=color, italic=italic)

def _border() -> Border:
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_left_border() -> Border:
    thick = Side(style="medium", color=C_ORANGE)
    thin  = Side(style="thin",   color=C_BORDER)
    return Border(left=thick, right=thin, top=thin, bottom=thin)

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _apply(cell, fill=None, font=None, alignment=None, border=None):
    if fill:      cell.fill      = fill
    if font:      cell.font      = font
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border

def _set_col_width(ws, col_letter: str, width: float):
    ws.column_dimensions[col_letter].width = width

def _row_height(ws, row: int, height: float):
    ws.row_dimensions[row].height = height

def _title_block(ws, title: str, subtitle: str):
    """Insere bloco de título nas linhas 1 e 2 e retorna próxima linha livre."""
    ws.merge_cells("A1:K1")
    ws["A1"] = title
    _apply(ws["A1"],
           fill=_fill(C_ORANGE),
           font=_font(16, bold=True, color=C_WHITE),
           alignment=_align("center"))
    _row_height(ws, 1, 36)

    ws.merge_cells("A2:K2")
    ws["A2"] = subtitle
    _apply(ws["A2"],
           fill=_fill(C_NAVY),
           font=_font(9, color=C_WHITE, italic=True),
           alignment=_align("center"))
    _row_height(ws, 2, 18)

    _row_height(ws, 3, 8)
    return 4


def _section_label(ws, row: int, text: str, ncols: int = 11):
    end_col = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{end_col}{row}")
    ws[f"A{row}"] = text
    _apply(ws[f"A{row}"],
           fill=_fill(C_LIGHT_OG),
           font=_font(10, bold=True, color=C_NAVY),
           border=_thick_left_border(),
           alignment=_align("left"))
    _row_height(ws, row, 22)


def _header_row(ws, row: int, headers: list[str]):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        _apply(cell,
               fill=_fill(C_NAVY),
               font=_font(10, bold=True, color=C_WHITE),
               border=_border(),
               alignment=_align("center"))
    _row_height(ws, row, 20)


def _data_cell(ws, row: int, col: int, value, alt: bool = False,
               fill_override: str | None = None,
               bold: bool = False,
               align_h: str = "left"):
    cell = ws.cell(row=row, column=col, value=value)
    bg = fill_override if fill_override else (C_ALT_ROW if alt else C_WHITE)
    _apply(cell,
           fill=_fill(bg),
           font=_font(10, bold=bold),
           border=_border(),
           alignment=_align(align_h, wrap=True))
    _row_height(ws, row, 18)
    return cell


# ══════════════════════════════════════════════════════════════════════════════
# COLETA DE DADOS
# ══════════════════════════════════════════════════════════════════════════════

_BRT = timezone(timedelta(hours=-3))

FONTES = [
    ("Porto de Santos – Esperados",  buscar_santos_esperados,  "Porto de Santos"),
    ("Porto de Santos – Atracados",  buscar_santos_atracados,  "Porto de Santos"),
    ("Porto de Itaqui",              buscar_itaqui,            "Porto de Itaqui"),
    ("Porto de Paranaguá",           buscar_paranagua,         "Porto de Paranaguá"),
    ("Porto de São Sebastião",       buscar_sao_sebastiao,     "Porto de São Sebastião"),
    ("Porto de Suape",               buscar_suape,             "Porto de Suape"),
]


def coletar_dados() -> tuple[pd.DataFrame, list[dict]]:
    """
    Executa todos os scrapers e retorna (DataFrame consolidado, log de coleta).
    O log contém uma entrada por fonte com: nome, status, registros, mensagem_de_erro.
    """
    log: list[dict] = []
    tabelas: list[pd.DataFrame] = []

    for nome, fn, _porto in FONTES:
        print(f"  Coletando {nome}... ", end="", flush=True)
        try:
            df = fn()
            n  = len(df)
            tabelas.append(df)
            log.append({"fonte": nome, "status": "OK", "registros": n, "erro": ""})
            print(f"{n} registro(s) ✓")
        except Exception as e:
            tabelas.append(pd.DataFrame())
            msg = str(e).split("\n")[0][:120]
            log.append({"fonte": nome, "status": "ERRO", "registros": 0, "erro": msg})
            print(f"ERRO — {msg}")

    resultado = consolidar(*tabelas)
    return resultado, log


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Resumo da Coleta
# ══════════════════════════════════════════════════════════════════════════════

def _build_resumo(wb: openpyxl.Workbook, df: pd.DataFrame,
                  log: list[dict], ts: datetime):
    ws = wb.active
    ws.title = "Resumo"

    agora_str = ts.strftime("%d/%m/%Y às %H:%M (BRT)")
    next_row = _title_block(
        ws,
        "NAVIOS DIESEL — RESUMO DA COLETA",
        f"Gerado em {agora_str}  |  Itaú BBA  |  Portos: Santos, Itaqui, Paranaguá, São Sebastião, Suape",
    )

    # ── Log de coleta ──────────────────────────────────────────────────────
    _section_label(ws, next_row, "LOG DE COLETA — STATUS POR FONTE", ncols=5)
    next_row += 1

    _header_row(ws, next_row, ["Fonte", "Status", "Registros", "Mensagem de Erro"])
    next_row += 1

    for i, entry in enumerate(log):
        alt  = i % 2 == 1
        ok   = entry["status"] == "OK"
        bg   = C_SUCCESS if ok else C_ERROR
        txt  = C_OK_DARK if ok else C_ERR_DARK

        cell_fonte = ws.cell(row=next_row, column=1, value=entry["fonte"])
        _apply(cell_fonte, fill=_fill(C_ALT_ROW if alt else C_WHITE),
               font=_font(10), border=_border(), alignment=_align("left"))

        cell_status = ws.cell(row=next_row, column=2, value=entry["status"])
        _apply(cell_status, fill=_fill(bg),
               font=_font(10, bold=True, color=txt),
               border=_border(), alignment=_align("center"))

        cell_regs = ws.cell(row=next_row, column=3, value=entry["registros"])
        _apply(cell_regs, fill=_fill(C_ALT_ROW if alt else C_WHITE),
               font=_font(10, bold=True), border=_border(), alignment=_align("center"))

        cell_err = ws.cell(row=next_row, column=4, value=entry["erro"])
        _apply(cell_err, fill=_fill(C_ALT_ROW if alt else C_WHITE),
               font=_font(9, italic=True), border=_border(), alignment=_align("left", wrap=True))
        _row_height(ws, next_row, 18)

        next_row += 1

    next_row += 1

    # ── Resumo por porto ───────────────────────────────────────────────────
    _section_label(ws, next_row, "RESUMO POR PORTO", ncols=5)
    next_row += 1

    _header_row(ws, next_row, ["Porto", "Status", "Qtd. Navios", "Volume Total (m³)", "% do Total"])
    total_vol_row = next_row  # guardamos para referência de fórmula
    next_row += 1

    if not df.empty:
        grp = (
            df[df["Status"] != "ERRO_COLETA"]
            .groupby(["Porto", "Status"], sort=False)
            .agg(navios=("Navio", "count"), volume=("Quantidade (m³)", "sum"))
            .reset_index()
            .sort_values(["Porto", "navios"], ascending=[True, False])
        )
        total_volume = grp["volume"].sum()

        first_data_row = next_row
        for i, row_data in grp.iterrows():
            alt  = (next_row - first_data_row) % 2 == 1
            bg   = STATUS_FILL.get(row_data["Status"], C_WHITE)

            _data_cell(ws, next_row, 1, row_data["Porto"],   alt=alt)
            _data_cell(ws, next_row, 2, row_data["Status"],  fill_override=bg, align_h="center")
            _data_cell(ws, next_row, 3, int(row_data["navios"]),  alt=alt, align_h="center", bold=True)
            vol = round(row_data["volume"], 0)
            _data_cell(ws, next_row, 4, vol, alt=alt, align_h="right", bold=True)
            ws.cell(row=next_row, column=4).number_format = "#,##0"
            pct = (row_data["volume"] / total_volume * 100) if total_volume else 0
            _data_cell(ws, next_row, 5, round(pct, 1), alt=alt, align_h="center")
            ws.cell(row=next_row, column=5).number_format = '0.0"%"'
            next_row += 1

        # Total
        total_row = next_row
        for c in range(1, 6):
            cell = ws.cell(row=total_row, column=c)
            cell.fill      = _fill(C_ORANGE)
            cell.font      = _font(10, bold=True, color=C_WHITE)
            cell.border    = _border()
            cell.alignment = _align("center")
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=3, value=int(grp["navios"].sum()))
        total_cell = ws.cell(row=total_row, column=4, value=round(total_volume, 0))
        total_cell.number_format = "#,##0"
        ws.cell(row=total_row, column=5, value="100.0%")
        _row_height(ws, total_row, 22)

    # ── Legenda de status ──────────────────────────────────────────────────
    next_row += 2
    _section_label(ws, next_row, "LEGENDA DE STATUS", ncols=5)
    next_row += 1

    legenda = [
        ("Atracado",   C_SUCCESS, "Navio está atracado no berço e descarregando"),
        ("Esperado",   C_WARN,    "Navio ainda não chegou — aguardando atracação"),
        ("Fundeado",   C_INFO,    "Navio ancorado fora do porto, aguardando berço"),
        ("Ao Largo",   C_INFO,    "Navio em posição de espera offshore"),
        ("Programado", C_WARN,    "Navio programado para chegada futura"),
        ("Despachado", C_MUTED,   "Navio já partiu após concluir operação"),
    ]
    for status, bg, descricao in legenda:
        cell_s = ws.cell(row=next_row, column=1, value=status)
        _apply(cell_s, fill=_fill(bg), font=_font(10, bold=True),
               border=_border(), alignment=_align("center"))
        cell_d = ws.cell(row=next_row, column=2, value=descricao)
        ws.merge_cells(f"B{next_row}:E{next_row}")
        _apply(cell_d, fill=_fill(C_WHITE), font=_font(10),
               border=_border(), alignment=_align("left"))
        _row_height(ws, next_row, 18)
        next_row += 1

    # ── Larguras das colunas ───────────────────────────────────────────────
    widths = {"A": 32, "B": 16, "C": 14, "D": 20, "E": 14}
    for col, w in widths.items():
        _set_col_width(ws, col, w)

    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Detalhes dos Navios
# ══════════════════════════════════════════════════════════════════════════════

def _fmt_date(val) -> str:
    """Formata data/hora para exibição legível."""
    if pd.isna(val) or str(val).strip() in ("", "nan", "NaT"):
        return ""
    try:
        dt = pd.to_datetime(str(val), dayfirst=True)
        return dt.strftime("%d/%m/%Y %H:%M") if dt.hour + dt.minute > 0 else dt.strftime("%d/%m/%Y")
    except Exception:
        return str(val)


def _fmt_vol(val) -> str | float:
    if pd.isna(val):
        return ""
    try:
        return round(float(val), 0)
    except Exception:
        return ""


def _build_detalhes(wb: openpyxl.Workbook, df: pd.DataFrame, ts: datetime):
    ws = wb.create_sheet("Detalhes dos Navios")

    agora_str = ts.strftime("%d/%m/%Y às %H:%M (BRT)")
    _title_block(
        ws,
        "NAVIOS DIESEL — DETALHES POR NAVIO",
        f"Gerado em {agora_str}  |  Todos os navios com Óleo Diesel nos portos monitorados",
    )

    headers = [
        "Porto", "Status", "Navio",
        "Volume (m³)", "Qtd. Original", "Unidade",
        "ETA / Chegada", "Início Descarga", "Fim Descarga",
        "Origem", "Terminal",
    ]
    header_row = 4
    _header_row(ws, header_row, headers)

    col_map = {
        "Porto":            "Porto",
        "Status":           "Status",
        "Navio":            "Navio",
        "Volume (m³)":      "Quantidade (m³)",
        "Qtd. Original":    "Quantidade Original",
        "Unidade":          "Unidade Origem",
        "ETA / Chegada":    "Chegada",
        "Início Descarga":  "Atracação",
        "Fim Descarga":     "Desatracação",
        "Origem":           "Origem",
        "Terminal":         "Terminal",
    }

    navios_display = df[df["Status"] != "ERRO_COLETA"].copy() if not df.empty else df

    for i, (_, row) in enumerate(navios_display.iterrows()):
        data_row = header_row + 1 + i
        alt      = i % 2 == 1
        status   = str(row.get("Status", "")).strip()
        row_bg   = STATUS_FILL.get(status, C_WHITE)

        for c, (col_label, src_col) in enumerate(col_map.items(), start=1):
            raw = row.get(src_col, "")

            if col_label == "Status":
                val    = status
                bg     = row_bg
                h_align = "center"
                is_bold = True
            elif col_label == "Volume (m³)":
                val    = _fmt_vol(raw)
                bg     = C_ALT_ROW if alt else C_WHITE
                h_align = "right"
                is_bold = True
            elif col_label in ("ETA / Chegada", "Início Descarga", "Fim Descarga"):
                val    = _fmt_date(raw)
                bg     = C_ALT_ROW if alt else C_WHITE
                h_align = "center"
                is_bold = False
            elif col_label in ("Qtd. Original",):
                val    = _fmt_vol(raw)
                bg     = C_ALT_ROW if alt else C_WHITE
                h_align = "right"
                is_bold = False
            else:
                val    = "" if pd.isna(raw) else str(raw)
                bg     = C_ALT_ROW if alt else C_WHITE
                h_align = "left"
                is_bold = False

            cell = ws.cell(row=data_row, column=c, value=val)
            _apply(cell,
                   fill=_fill(bg),
                   font=_font(10, bold=is_bold),
                   border=_border(),
                   alignment=_align(h_align))

            if col_label == "Volume (m³)" and isinstance(val, (int, float)):
                cell.number_format = "#,##0"

        _row_height(ws, data_row, 18)

    # Linha de total no final
    if not navios_display.empty:
        total_row = header_row + 1 + len(navios_display)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row, column=c)
            cell.fill   = _fill(C_ORANGE)
            cell.font   = _font(10, bold=True, color=C_WHITE)
            cell.border = _border()
            cell.alignment = _align("center")
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=3,
                value=f"{len(navios_display)} navio(s)")
        vol_total = navios_display["Quantidade (m³)"].dropna().sum()
        cell_tot  = ws.cell(row=total_row, column=4, value=round(vol_total, 0))
        cell_tot.number_format = "#,##0"
        _row_height(ws, total_row, 22)

    # Larguras das colunas
    widths = {
        "A": 24, "B": 14, "C": 28, "D": 14,
        "E": 14, "F": 10, "G": 18, "H": 18,
        "I": 18, "J": 18, "K": 20,
    }
    for col, w in widths.items():
        _set_col_width(ws, col, w)

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Por Porto (uma seção por porto)
# ══════════════════════════════════════════════════════════════════════════════

def _build_por_porto(wb: openpyxl.Workbook, df: pd.DataFrame, ts: datetime):
    ws = wb.create_sheet("Por Porto")

    agora_str = ts.strftime("%d/%m/%Y às %H:%M (BRT)")
    _title_block(
        ws,
        "NAVIOS DIESEL — VISÃO POR PORTO",
        f"Gerado em {agora_str}  |  Cada seção detalha os navios de um porto",
    )

    headers = ["Status", "Navio", "Volume (m³)", "ETA / Chegada",
               "Início Descarga", "Fim Descarga", "Origem", "Terminal"]

    current_row = 4
    navios_display = df[df["Status"] != "ERRO_COLETA"].copy() if not df.empty else df
    portos = navios_display["Porto"].unique() if not navios_display.empty else []

    for porto in portos:
        porto_df = navios_display[navios_display["Porto"] == porto].copy()

        # Cabeçalho do porto
        ws.merge_cells(f"A{current_row}:I{current_row}")
        cell_porto = ws.cell(row=current_row, column=1, value=porto.upper())
        _apply(cell_porto,
               fill=_fill(C_ORANGE),
               font=_font(12, bold=True, color=C_WHITE),
               alignment=_align("center"))
        _row_height(ws, current_row, 26)

        # Sub-cabeçalho com volume total do porto
        vol_porto = porto_df["Quantidade (m³)"].dropna().sum()
        current_row += 1
        ws.merge_cells(f"A{current_row}:I{current_row}")
        cell_sub = ws.cell(
            row=current_row, column=1,
            value=(f"{len(porto_df)} navio(s)   |   "
                   f"Volume total: {vol_porto:,.0f} m³"),
        )
        _apply(cell_sub,
               fill=_fill(C_NAVY),
               font=_font(10, color=C_WHITE, italic=True),
               alignment=_align("center"))
        _row_height(ws, current_row, 18)

        current_row += 1
        _header_row(ws, current_row, headers)
        col_map = {
            "Status":          "Status",
            "Navio":           "Navio",
            "Volume (m³)":     "Quantidade (m³)",
            "ETA / Chegada":   "Chegada",
            "Início Descarga": "Atracação",
            "Fim Descarga":    "Desatracação",
            "Origem":          "Origem",
            "Terminal":        "Terminal",
        }
        current_row += 1

        for i, (_, row) in enumerate(porto_df.iterrows()):
            alt    = i % 2 == 1
            status = str(row.get("Status", "")).strip()

            for c, (col_label, src_col) in enumerate(col_map.items(), start=1):
                raw = row.get(src_col, "")

                if col_label == "Status":
                    val    = status
                    bg     = STATUS_FILL.get(status, C_WHITE)
                    h_align = "center"
                    is_bold = True
                elif col_label == "Volume (m³)":
                    val    = _fmt_vol(raw)
                    bg     = C_ALT_ROW if alt else C_WHITE
                    h_align = "right"
                    is_bold = True
                elif col_label in ("ETA / Chegada", "Início Descarga", "Fim Descarga"):
                    val    = _fmt_date(raw)
                    bg     = C_ALT_ROW if alt else C_WHITE
                    h_align = "center"
                    is_bold = False
                else:
                    val    = "" if pd.isna(raw) else str(raw)
                    bg     = C_ALT_ROW if alt else C_WHITE
                    h_align = "left"
                    is_bold = False

                cell = ws.cell(row=current_row, column=c, value=val)
                _apply(cell,
                       fill=_fill(bg),
                       font=_font(10, bold=is_bold),
                       border=_border(),
                       alignment=_align(h_align))
                if col_label == "Volume (m³)" and isinstance(val, (int, float)):
                    cell.number_format = "#,##0"

            _row_height(ws, current_row, 18)
            current_row += 1

        current_row += 2  # espaço entre portos

    # Larguras das colunas
    widths = {
        "A": 14, "B": 32, "C": 14, "D": 18,
        "E": 18, "F": 18, "G": 22, "H": 22, "I": 10,
    }
    for col, w in widths.items():
        _set_col_width(ws, col, w)

    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ts = datetime.now(_BRT)
    print()
    print("══════════════════════════════════════════════════════════")
    print("  MONITORAMENTO NAVIOS DIESEL — EXPORTAÇÃO PARA EXCEL")
    print(f"  {ts.strftime('%d/%m/%Y %H:%M')} BRT")
    print("══════════════════════════════════════════════════════════")
    print()

    print("[ 1/3 ] Coletando dados dos portos...")
    df, log = coletar_dados()

    n_ok   = sum(1 for e in log if e["status"] == "OK")
    n_err  = sum(1 for e in log if e["status"] == "ERRO")
    total  = len(df[df["Status"] != "ERRO_COLETA"]) if not df.empty else 0
    print(f"\n  Fontes: {n_ok} OK / {n_err} com erro")
    print(f"  Navios encontrados: {total}")

    print("\n[ 2/3 ] Gerando Excel...")
    wb = openpyxl.Workbook()
    _build_resumo(wb, df, log, ts)
    _build_detalhes(wb, df, ts)
    _build_por_porto(wb, df, ts)

    # Definir propriedades do documento
    wb.properties.title   = "Navios Diesel — Monitoramento Itaú BBA"
    wb.properties.creator = "navios_esperados.py / exportar_para_excel.py"

    print("\n[ 3/3 ] Salvando arquivo...")
    nome_arquivo = f"Navios_Diesel_{ts.strftime('%Y%m%d_%H%M')}.xlsx"
    pasta        = Path(__file__).parent
    caminho      = pasta / nome_arquivo
    wb.save(caminho)

    print()
    print("══════════════════════════════════════════════════════════")
    print(f"  ✓ Arquivo salvo em:")
    print(f"    {caminho}")
    print()
    print("  Abas do Excel:")
    print("    • Resumo            → log de coleta + resumo por porto")
    print("    • Detalhes dos Navios → todos os navios com volume e datas")
    print("    • Por Porto         → uma seção por porto")
    print("══════════════════════════════════════════════════════════")
    print()


if __name__ == "__main__":
    main()
