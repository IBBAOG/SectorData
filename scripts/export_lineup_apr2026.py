"""Export the Apr-2026 lineup (Delivered + Pending) for the 2026-04-23 01:19 BRT snapshot."""
import os
import sys
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
ENV_FILE = ROOT / ".env"
for line in ENV_FILE.read_text().splitlines():
    if line.startswith("#") or "=" not in line:
        continue
    k, v = line.split("=", 1)
    os.environ.setdefault(k.strip(), v.strip())

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_KEY"]

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
}

SNAPSHOT = "2026-04-23T13:19:00"
TARGET_MONTH = "2026-04"
OUTPUT = ROOT / "output" / f"lineup_{TARGET_MONTH}.xlsx"


def rpc(fn: str, params: dict) -> list | dict:
    r = requests.post(f"{SUPABASE_URL}/rest/v1/rpc/{fn}", headers=HEADERS, json=params, timeout=60)
    r.raise_for_status()
    return r.json()


def main() -> int:
    coletas = rpc("get_nd_coletas_distintas", {})
    if SNAPSHOT not in coletas:
        matches = [c for c in coletas if c.startswith("2026-04-23")]
        print(f"Exact snapshot {SNAPSHOT} not found. 2026-04-23 snapshots: {matches}", file=sys.stderr)
        return 1
    print(f"Using snapshot {SNAPSHOT} BRT")

    volume = rpc("get_nd_volume_mensal_descarga", {"p_collected_at": SNAPSHOT})
    apr = next((r for r in volume if r["month"] == TARGET_MONTH), None)
    if not apr:
        print(f"No {TARGET_MONTH} row in volume RPC", file=sys.stderr)
        return 1
    total_expected = round(apr["discharged_volume"] + apr["pending_volume"] + apr["indeterminate_volume"])
    print(f"Apr 2026 chart totals — Discharged: {apr['discharged_volume']:,.0f}  Pending: {apr['pending_volume']:,.0f}  Indet: {apr['indeterminate_volume']:,.0f}  TOTAL: {total_expected:,}")

    descarregados = rpc("get_nd_navios_descarregados", {"p_collected_at": SNAPSHOT})
    # Chart buckets Discharged by `last_seen` month (not the ETA-preferred discharge_month)
    delivered_apr = [r for r in descarregados if (r.get("last_seen") or "")[:7] == TARGET_MONTH]

    navios = rpc("get_nd_navios", {"p_collected_at": SNAPSHOT})
    active = [n for n in navios if n["status"] not in ("Despachado", "ERRO_COLETA")]

    def discharge_month(n: dict) -> str:
        raw = n.get("eta") or n.get("inicio_descarga") or n.get("fim_descarga") or n.get("collected_at")
        if not raw:
            return ""
        return raw[:7]

    pending_apr = [n for n in active if discharge_month(n) == TARGET_MONTH]

    delivered_sum = sum(r["last_volume"] or 0 for r in delivered_apr)
    pending_sum = sum(n.get("quantidade_convertida") or 0 for n in pending_apr)
    print(f"Per-vessel sums — Delivered: {delivered_sum:,.0f} ({len(delivered_apr)} vessels)  Pending: {pending_sum:,.0f} ({len(pending_apr)} vessels)  TOTAL: {delivered_sum + pending_sum:,.0f}")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="000512")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    section_fill_d = PatternFill("solid", fgColor="000000")
    section_fill_p = PatternFill("solid", fgColor="FF5000")
    section_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    total_font = Font(bold=True, name="Arial", size=11)
    total_fill = PatternFill("solid", fgColor="FFF3CD")

    ws = wb.active
    ws.title = "Apr 2026 Lineup"

    def banner(row: int, text: str, fill: PatternFill, span: int):
        ws.cell(row=row, column=1, value=text).font = section_font
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)

    headers_d = ["Category", "Port", "Vessel", "Volume (m3)", "Last Seen (BRT)", "Discharge Month"]
    headers_p = ["Category", "Port", "Status", "Vessel", "Volume (m3)", "ETA", "Unload Start", "Unload End", "Origin", "Flag", "IMO", "MMSI"]
    span = max(len(headers_d), len(headers_p))

    banner(1, "DISCHARGED / DELIVERED VESSELS  —  April 2026 (snapshot 2026-04-23 01:19 BRT)", section_fill_d, span)

    for i, h in enumerate(headers_d, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    row = 3
    for r in sorted(delivered_apr, key=lambda x: (x["porto"], -float(x["last_volume"] or 0))):
        ws.cell(row=row, column=1, value="Discharged")
        ws.cell(row=row, column=2, value=(r["porto"] or "").replace("Porto de ", ""))
        ws.cell(row=row, column=3, value=r["navio"])
        ws.cell(row=row, column=4, value=r["last_volume"]).number_format = "#,##0"
        ws.cell(row=row, column=5, value=r["last_seen"])
        ws.cell(row=row, column=6, value=r["discharge_month"])
        row += 1

    ws.cell(row=row, column=3, value="Subtotal — Discharged").font = total_font
    ws.cell(row=row, column=3).fill = total_fill
    ws.cell(row=row, column=4, value=delivered_sum).number_format = "#,##0"
    ws.cell(row=row, column=4).font = total_font
    ws.cell(row=row, column=4).fill = total_fill
    row += 2

    banner(row, "PENDING DISCHARGE / EXPECTED VESSELS  —  April 2026", section_fill_p, span)
    row += 1
    for i, h in enumerate(headers_p, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    for n in sorted(pending_apr, key=lambda x: (x["porto"], x.get("status") or "", -(float(x.get("quantidade_convertida") or 0)))):
        ws.cell(row=row, column=1, value="Pending")
        ws.cell(row=row, column=2, value=(n["porto"] or "").replace("Porto de ", ""))
        ws.cell(row=row, column=3, value=n.get("status"))
        ws.cell(row=row, column=4, value=n.get("navio"))
        ws.cell(row=row, column=5, value=n.get("quantidade_convertida")).number_format = "#,##0"
        ws.cell(row=row, column=6, value=(n.get("eta") or "")[:16].replace("T", " "))
        ws.cell(row=row, column=7, value=(n.get("inicio_descarga") or "")[:16].replace("T", " "))
        ws.cell(row=row, column=8, value=(n.get("fim_descarga") or "")[:16].replace("T", " "))
        ws.cell(row=row, column=9, value=n.get("origem"))
        ws.cell(row=row, column=10, value=n.get("flag"))
        ws.cell(row=row, column=11, value=n.get("imo"))
        ws.cell(row=row, column=12, value=n.get("mmsi"))
        row += 1

    ws.cell(row=row, column=4, value="Subtotal — Pending").font = total_font
    ws.cell(row=row, column=4).fill = total_fill
    ws.cell(row=row, column=5, value=pending_sum).number_format = "#,##0"
    ws.cell(row=row, column=5).font = total_font
    ws.cell(row=row, column=5).fill = total_fill
    row += 2

    ws.cell(row=row, column=3, value="TOTAL — April 2026 (should match chart = 1,432,493)").font = total_font
    ws.cell(row=row, column=3).fill = total_fill
    ws.cell(row=row, column=4, value=delivered_sum + pending_sum).number_format = "#,##0"
    ws.cell(row=row, column=4).font = total_font
    ws.cell(row=row, column=4).fill = total_fill

    widths = [14, 18, 18, 32, 18, 22, 18, 18, 22, 14, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")
    print(f"Grand total: {delivered_sum + pending_sum:,.0f}  (expected 1,432,493 — delta {(delivered_sum + pending_sum) - 1_432_493:+,.0f})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
