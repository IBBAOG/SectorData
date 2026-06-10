"""
[DEPRECATED] Generate a single-metric Stock Guide scenario-grid Excel TEMPLATE.

⚠️  DEPRECATED (2026-06-10). The canonical template is now DOWNLOADED FROM THE
    ADMIN PANEL ("Download template" button on the scenario-grid shell, generated
    in-browser). The admin template is the multi-metric, multi-sheet v2 contract
    (one sheet per output metric, coordinates read positionally) that the uploader
    `stock_guide_brent_grid_upload.py` expects.

    This script remains ONLY as a single-metric / offline fallback that emits ONE
    sheet (named `target_price`, the legacy default output). It REFUSES to run
    against a shell configured with multiple output metrics — use the Admin Panel
    "Download template" button for those. Prefer the Admin button in all cases.

────────────────────────────────────────────────────────────────────────────
Generate the (single-metric) Stock Guide multi-axis scenario-grid Excel TEMPLATE
(LONG format).

The analyst fills, per company, a dense Cartesian mesh of (driver levels ->
target price) points across 1, 2 or 3 driver axes (e.g. Avg Brent 2026 / 2027 /
2028+). Every row is one model run = one scenario. The frontend (/stock-guide)
interpolates this mesh multilinearly live against the current driver levels. The
filled grid is uploaded with `scripts/manual/stock_guide_brent_grid_upload.py`.

This script only produces the skeleton the analyst fills in:

    - LONG format, single sheet.
    - One coordinate column per axis, header = the EXACT driver_key
      (e.g. avg_brent_2026), pre-filled with the full Cartesian product of the
      per-axis levels (the FIRST axis varies slowest).
    - One column per ticker, header = ticker, cells LEFT BLANK (the analyst types
      the target price R$/share for each scenario).

Axes are resolved (in priority order):
    1. --axes avg_brent_2026,avg_brent_2027,avg_brent_2028   (offline, explicit)
    2. --sensitivity-id N / --table-title "..."  → read definition.grid.axes via
       the ANON-callable RPC get_stock_guide_sensitivity_tables (returns the
       shell's `definition`, grid block intact).
    3. Nothing passed → default 3 Brent axes (avg_brent_2026/2027/2028).

Per-axis ranges via the repeatable flag:
    --range KEY=MIN:MAX:STEP    (e.g. --range avg_brent_2026=40:150:10)
Any axis without an explicit --range uses the default 40:150:10 (12 levels).
Levels are rounded to 6 decimals (same as the uploader — neutralizes float drift).

Tickers are discovered from `stock_guide_companies` (visible tickers, ordered by
display_order) via the public hide-aware RPC `get_stock_guide_comps` using the
ANON key (read-only, RLS-respecting). Override with --tickers.

The output is the EXACT file the uploader reads by default:
    C:\\Users\\eduar\\dashboard_projeto\\data\\stock_guide_brent_grid.xlsx
(override with --out). That path is gitignored — never committed.

────────────────────────────────────────────────────────────────────────────
Usage:
    # Default: 3 Brent axes (avg_brent_2026/2027/2028), 40->150 step 10 each,
    # tickers from Supabase:
    python scripts/manual/make_brent_grid_template.py

    # Pull axes from a shell + custom ranges:
    python scripts/manual/make_brent_grid_template.py --sensitivity-id 7 \
        --range avg_brent_2026=40:120:10 --range avg_brent_2027=40:120:10 --force

    # Offline: explicit axes + ranges + tickers, custom output:
    python scripts/manual/make_brent_grid_template.py \
        --axes avg_brent_2026,avg_brent_2027 \
        --range avg_brent_2026=50:120:10 --range avg_brent_2027=50:120:10 \
        --tickers PETR4,PRIO3,RECV3 --out C:\\path\\to\\grid.xlsx --force

Credentials for axis lookup (sensitivity RPC) and ticker discovery (env vars,
fall back to .env / .env.local):
    NEXT_PUBLIC_SUPABASE_URL  (or SUPABASE_URL)
    NEXT_PUBLIC_SUPABASE_ANON_KEY  (or SUPABASE_ANON_KEY)
"""

from __future__ import annotations

import argparse
import os
import sys
from itertools import product
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


# ── Defaults ────────────────────────────────────────────────────────────────────

_DEFAULT_OUT = r"C:\Users\eduar\dashboard_projeto\data\stock_guide_brent_grid.xlsx"
_DEFAULT_AXES = ["avg_brent_2026", "avg_brent_2027", "avg_brent_2028"]
_DEFAULT_RANGE = (40.0, 150.0, 10.0)   # min:max:step -> 12 levels
_COORD_ROUND = 6
_WARN_ROWS = 60_000


# ── Env / credentials ──────────────────────────────────────────────────────────

def _load_env_files() -> dict[str, str]:
    """Merge .env then .env.local (latter wins) from the current working dir."""
    result: dict[str, str] = {}
    for fname in (".env", ".env.local"):
        p = Path(fname)
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            result[key.strip()] = val.strip()
    return result


def _get_anon_credentials() -> tuple[str, str]:
    env = _load_env_files()
    url = (
        os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
        or os.environ.get("SUPABASE_URL")
        or env.get("NEXT_PUBLIC_SUPABASE_URL")
        or env.get("SUPABASE_URL")
        or ""
    )
    key = (
        os.environ.get("NEXT_PUBLIC_SUPABASE_ANON_KEY")
        or os.environ.get("SUPABASE_ANON_KEY")
        or env.get("NEXT_PUBLIC_SUPABASE_ANON_KEY")
        or env.get("SUPABASE_ANON_KEY")
        or ""
    )
    return url, key


def _make_anon_client():
    url, key = _get_anon_credentials()
    if not url or not key:
        return None
    try:
        from supabase import create_client
    except ImportError:
        print("ERROR: supabase-py not installed (pip install supabase).")
        sys.exit(1)
    return create_client(url, key)


# ── Axis resolution ──────────────────────────────────────────────────────────────

def _axes_from_shell(sb, args: argparse.Namespace) -> list[str] | None:
    """Read definition.grid.axes from a shell via the anon RPC
    get_stock_guide_sensitivity_tables (returns definition with grid intact).
    Returns the ordered list of driver_keys, or exits with a clear error."""
    if sb is None:
        print(
            "ERROR: --sensitivity-id / --table-title needs anon Supabase "
            "credentials (NEXT_PUBLIC_SUPABASE_URL / NEXT_PUBLIC_SUPABASE_ANON_KEY). "
            "Set env vars / .env.local, or pass --axes explicitly (offline)."
        )
        sys.exit(1)
    try:
        resp = sb.rpc("get_stock_guide_sensitivity_tables", {}).execute()
    except Exception as e:
        print(f"ERROR: get_stock_guide_sensitivity_tables RPC failed: {e}")
        sys.exit(1)

    rows = resp.data or []
    target = None
    if args.sensitivity_id is not None:
        for r in rows:
            if int(r.get("id", -1)) == int(args.sensitivity_id):
                target = r
                break
        if target is None:
            print(
                f"ERROR: no sensitivity table visible with id={args.sensitivity_id} "
                "(anon caller). Check the id, or pass --axes."
            )
            sys.exit(1)
    else:
        matches = [r for r in rows if str(r.get("title")) == args.table_title]
        if len(matches) == 0:
            print(
                f"ERROR: no sensitivity table with title={args.table_title!r} "
                "(anon caller). Check the title, or pass --axes."
            )
            sys.exit(1)
        if len(matches) > 1:
            ids = ", ".join(str(r.get("id")) for r in matches)
            print(
                f"ERROR: title {args.table_title!r} matches {len(matches)} tables "
                f"(ids: {ids}). Pass --sensitivity-id."
            )
            sys.exit(1)
        target = matches[0]

    definition = target.get("definition")
    if not isinstance(definition, dict):
        print("ERROR: shell has no JSON definition. Re-save the shell in the Admin Panel.")
        sys.exit(1)
    grid = definition.get("grid")
    if not isinstance(grid, dict):
        print(
            "ERROR: shell definition has no `grid` block — not a scenario-grid "
            "table. Re-save the shell in the Admin Panel."
        )
        sys.exit(1)
    axes = grid.get("axes")
    if not isinstance(axes, list) or not axes:
        if "x_driver_key" in grid:
            print(
                "ERROR: shell carries the legacy 1-D grid shape "
                "({x_driver_key,...}). Re-save the shell in the Admin Panel."
            )
            sys.exit(1)
        print("ERROR: shell definition.grid has no `axes` list. Re-save the shell.")
        sys.exit(1)
    keys = [str(a.get("driver_key") or a.get("driver_id") or "").strip()
            for a in axes if isinstance(a, dict)]
    keys = [k for k in keys if k]
    if not keys:
        print("ERROR: shell grid.axes has no valid driver_key. Re-save the shell.")
        sys.exit(1)
    if len(keys) > 3:
        print(f"ERROR: shell defines {len(keys)} axes — max 3. Re-save the shell.")
        sys.exit(1)

    # DEPRECATION GUARD: this fallback only emits ONE (target_price) sheet. If the
    # shell configures multiple output metrics, the analyst MUST use the Admin
    # Panel "Download template" button (multi-sheet v2). Exit informatively.
    raw_outputs = grid.get("outputs")
    if isinstance(raw_outputs, list):
        outs = [str(o).strip() for o in raw_outputs if str(o).strip()]
        if len(outs) > 1:
            print(
                f"\nThis shell is MULTI-METRIC (outputs={outs}). This deprecated "
                "generator only emits a single `target_price` sheet.\n"
                "→ Use the Admin Panel \"Download template\" button on the "
                "scenario-grid shell to get the multi-sheet v2 template, then "
                "upload it with stock_guide_brent_grid_upload.py.\n"
            )
            sys.exit(2)

    print(f"Axes (from shell): {keys}")
    return keys


def _resolve_axes(sb, args: argparse.Namespace) -> list[str]:
    if args.axes:
        keys = [a.strip() for a in args.axes.split(",") if a.strip()]
        if not keys:
            print("ERROR: --axes was empty after parsing.")
            sys.exit(1)
        if len(keys) > 3:
            print(f"ERROR: --axes has {len(keys)} axes — max 3.")
            sys.exit(1)
        if len(set(k.lower() for k in keys)) != len(keys):
            print(f"ERROR: --axes has duplicate driver keys: {keys}.")
            sys.exit(1)
        print(f"Axes (from --axes): {keys}")
        return keys
    if args.sensitivity_id is not None or args.table_title is not None:
        return _axes_from_shell(sb, args)
    print(f"Axes (default 3 Brent): {_DEFAULT_AXES}")
    return list(_DEFAULT_AXES)


# ── Range parsing + level building ───────────────────────────────────────────────

def _parse_ranges(raw: list[str] | None) -> dict[str, tuple[float, float, float]]:
    out: dict[str, tuple[float, float, float]] = {}
    for item in raw or []:
        if "=" not in item:
            print(f"ERROR: --range must be KEY=MIN:MAX:STEP (got {item!r}).")
            sys.exit(1)
        key, _, spec = item.partition("=")
        key = key.strip()
        parts = spec.split(":")
        if len(parts) != 3:
            print(f"ERROR: --range {item!r}: spec must be MIN:MAX:STEP.")
            sys.exit(1)
        try:
            lo, hi, step = (float(parts[0]), float(parts[1]), float(parts[2]))
        except ValueError:
            print(f"ERROR: --range {item!r}: MIN/MAX/STEP must be numeric.")
            sys.exit(1)
        out[key.lower()] = (lo, hi, step)
    return out


def _build_levels(lo: float, hi: float, step: float, key: str) -> list[float]:
    if step <= 0:
        print(f"ERROR: range for {key!r}: step must be > 0 (got {step}).")
        sys.exit(1)
    if hi < lo:
        print(f"ERROR: range for {key!r}: max ({hi}) must be >= min ({lo}).")
        sys.exit(1)
    levels: list[float] = []
    v = lo
    eps = step * 1e-9  # tiny epsilon so the upper bound is included despite float drift
    while v <= hi + eps:
        levels.append(round(v, _COORD_ROUND))
        v += step
    return levels


def _as_number(x: float):
    """Render whole numbers as int so Excel shows 40 not 40.0."""
    return int(x) if float(x).is_integer() else float(x)


# ── Ticker discovery ────────────────────────────────────────────────────────────

def _discover_tickers(sb) -> list[str]:
    """Visible tickers from stock_guide_companies, ordered by display_order, via
    the public hide-aware RPC get_stock_guide_comps (anon key)."""
    if sb is None:
        print(
            "ERROR: Could not find anon Supabase credentials "
            "(NEXT_PUBLIC_SUPABASE_URL / NEXT_PUBLIC_SUPABASE_ANON_KEY). "
            "Set env vars / .env.local, or pass --tickers explicitly."
        )
        sys.exit(1)
    try:
        resp = sb.rpc("get_stock_guide_comps", {}).execute()
    except Exception as e:
        print(f"ERROR: get_stock_guide_comps RPC failed: {e}")
        sys.exit(1)

    rows = resp.data or []
    if not rows:
        print(
            "ERROR: get_stock_guide_comps returned 0 rows. The anon caller may "
            "see no visible companies. Pass --tickers explicitly."
        )
        sys.exit(1)

    def _order_key(r: dict):
        d = r.get("display_order")
        return (d is None, d if d is not None else 0, str(r.get("ticker") or ""))

    rows_sorted = sorted(rows, key=_order_key)
    seen: set[str] = set()
    out: list[str] = []
    for r in rows_sorted:
        t = str(r.get("ticker") or "").strip()
        if t and t not in seen:
            seen.add(t)
            out.append(t)
    return out


# ── CLI ───────────────────────────────────────────────────────────────────────

def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description=(
            "Generate the EMPTY Stock Guide multi-axis scenario-grid Excel "
            "template (LONG: one coordinate column per axis named by its "
            "driver_key + one blank column per ticker). The analyst fills target "
            "prices; upload with stock_guide_brent_grid_upload.py."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--out", default=_DEFAULT_OUT, metavar="PATH",
                   help=f"Output .xlsx path (default: {_DEFAULT_OUT}).")
    src = p.add_mutually_exclusive_group()
    src.add_argument("--sensitivity-id", type=int, default=None, metavar="N",
                     help="Read axes from this shell's definition.grid.axes (anon RPC).")
    src.add_argument("--table-title", type=str, default=None, metavar="TITLE",
                     help="Read axes from the shell with this title (must be unique).")
    p.add_argument("--axes", type=str, default=None, metavar="K1,K2,...",
                   help="Explicit axis driver_keys (offline). Overrides shell lookup. "
                        "Default if nothing passed: avg_brent_2026,avg_brent_2027,avg_brent_2028.")
    p.add_argument("--range", action="append", default=None, dest="ranges",
                   metavar="KEY=MIN:MAX:STEP",
                   help="Per-axis level range (repeatable). e.g. "
                        "--range avg_brent_2026=40:150:10. Axes without an explicit "
                        f"range use {int(_DEFAULT_RANGE[0])}:{int(_DEFAULT_RANGE[1])}:{int(_DEFAULT_RANGE[2])}.")
    p.add_argument("--tickers", type=str, default=None, metavar="T1,T2,...",
                   help="Comma-separated ticker columns. Default: discover from "
                        "stock_guide_companies (visible, display_order).")
    p.add_argument("--force", action="store_true",
                   help="Overwrite the output file if it already exists "
                        "(DESTRUCTIVE — wipes any analyst input).")
    return p.parse_args()


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    args = _parse_args()

    print(
        "⚠️  DEPRECATED: prefer the Admin Panel \"Download template\" button "
        "(multi-metric v2).\n"
        "    This generator emits only a single `target_price` sheet "
        "(single-metric / offline fallback).\n"
    )

    out_path = Path(args.out)
    if out_path.exists() and not args.force:
        print(
            f"ERROR: Output already exists: {out_path}\n"
            "Refusing to overwrite (it may contain analyst input). "
            "Pass --force to regenerate from scratch."
        )
        sys.exit(1)

    # Build the anon client once (reused for axes + tickers); may be None offline.
    sb = _make_anon_client()

    # Resolve axes (explicit / shell / default)
    axis_keys = _resolve_axes(sb, args)
    dim = len(axis_keys)

    # Resolve per-axis ranges
    ranges = _parse_ranges(args.ranges)
    per_axis_levels: list[list[float]] = []
    for k in axis_keys:
        lo, hi, step = ranges.get(k.lower(), _DEFAULT_RANGE)
        levels = _build_levels(lo, hi, step, k)
        per_axis_levels.append(levels)
        print(
            f"  {k}: {_as_number(lo)} -> {_as_number(hi)} step {_as_number(step)} "
            f"({len(levels)} levels)"
        )

    # Resolve tickers
    if args.tickers:
        tickers = [t.strip() for t in args.tickers.split(",") if t.strip()]
        if not tickers:
            print("ERROR: --tickers was empty after parsing.")
            sys.exit(1)
        print(f"Tickers (from --tickers, {len(tickers)}): {tickers}")
    else:
        tickers = _discover_tickers(sb)
        print(f"Tickers (from stock_guide_companies, {len(tickers)}): {tickers}")

    # Cartesian product — first axis varies slowest (itertools.product does this).
    combos = list(product(*per_axis_levels))
    n_combos = len(combos)
    n_points = n_combos * len(tickers)

    # Build workbook
    wb = Workbook()
    ws = wb.active
    # Sheet name = metric key, so the v2 uploader maps this sheet to the
    # `target_price` output. (Single-metric fallback only.)
    ws.title = "target_price"

    headers = list(axis_keys) + tickers
    ws.append(headers)
    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = bold

    blanks = [None] * len(tickers)
    for combo in combos:
        ws.append([_as_number(c) for c in combo] + blanks)

    # Cosmetic: widen the coordinate columns; freeze header row.
    for ci in range(dim):
        ws.column_dimensions[ws.cell(row=1, column=ci + 1).column_letter].width = 16
    ws.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    dims_txt = " × ".join(str(len(lv)) for lv in per_axis_levels)
    print(f"\nWrote template: {out_path}")
    print(f"  Sheet: {ws.title!r}")
    print(f"  Coordinate columns ({dim}): {axis_keys}")
    print(f"  Ticker columns ({len(tickers)}): {tickers}")
    print(
        f"  {dims_txt} = {n_combos:,} scenarios × {len(tickers)} tickers "
        f"= {n_points:,} mesh points (ticker cells empty — analyst fills "
        "target prices R$/share)"
    )
    if n_points > _WARN_ROWS:
        print(
            f"  WARNING: {n_points:,} mesh points is large (> {_WARN_ROWS:,}). "
            "Keep ≤15 levels/axis for 3-D meshes (≤40×40 for 2-D)."
        )


if __name__ == "__main__":
    main()
